```python
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import re
import csv
from openpyxl.styles import Border, Side

def folder_name():
    fileList = []
    # 文件夹路径
    dataPath = ('.\\data')
    # 遍历各表文件夹名称
    dataItems = os.listdir(path=dataPath)
    # 获取所有文件夹下的文件名
    for dataItem in dataItems:
        dataItemPath = os.path.join(dataPath, dataItem)
        items = os.listdir(path=dataItemPath)
        for item in items:
            resultDataItemPath = os.path.join(dataItemPath, item)
            fileList.append(resultDataItemPath)
    return fileList


def business(fileList):

    dataLists = []
    for fileName in fileList:
        storeName = re.search(r'data\\(.*)\\', fileName).group(1)
        if re.search('(Business.*)', fileName) is not None:
            # 读表
            f = open(fileName, 'r')
            names = []
            with f:
                reader = csv.reader(f)
                for i, row in enumerate(reader):
                    if i >= 1:
                        asin = row[1]
                        visitor = int(re.sub(',','',row[3]))
                        order = int(re.sub(',','',row[8]))
                        sale1 = row[12]
                        if '￡' in sale1:
                            sale = float(re.sub(',', '', sale1[1:]))

                        elif sale1[0] == 'U':
                            sale = float(re.sub(',','',sale1[3:]))
                        else:
                            sale = float(re.sub(',', '', sale1))


                        dataList = [storeName, asin, visitor, order, sale]
                        if asin not in names:
                            names.append(asin)
                            dataLists.append(dataList)
                        else:
                            for data in dataLists:
                                if data[0] == dataList[0] and data[1] == dataList[1]:
                                    data[2] += visitor
                                    data[3] += order
                                    data[4] += sale
    return dataLists


def campaigns(fileList):
    resultAdvertisings = []
    advertisingDataLists = []
    campaignsDataLists = []
    notInCampaigns = []
    inCampaigns = []
    ads = []
    # 输出文件路径

    for fileName in fileList:
        storeName = re.search(r'data\\(.*)\\', fileName).group(1)
        if re.search('(Campaigns.*)', fileName) is not None:
            # 读表
            f = open(fileName, 'r')
            with f:
                reader = csv.reader(f)
                for i, row in enumerate(reader):
                    if i >= 1:
                        campaigns = row[1]
                        c_type = row[3]
                        cost = row[15]
                        advertisingLists = [campaigns, c_type, cost]
                        advertisingDataLists.append(advertisingLists)

    for rowData in workingSheetGG.iter_rows(min_row=2, max_col=3):
        campaigns = rowData[0].value
        sku = rowData[1].value
        campaignsList = [campaigns, sku]
        inCampaigns.append(campaigns)
        campaignsDataLists.append(campaignsList)
    for advertisingDataList in advertisingDataLists:
        if advertisingDataList[0] not in inCampaigns:
            notInCampaigns.append(advertisingDataList[0])
        else:
            for campaignsDataList in campaignsDataLists:
                if advertisingDataList[0] == campaignsDataList[0]:
                    advertisingDataList.append(campaignsDataList[1])
    for notInCampaign in notInCampaigns:
        r = [notInCampaign, '请填写SKU']
        workingSheetGG.append(r)
        # 保存工作表
        # workingbook.save(filename=workExcelPath)

    for advertisingDataList in advertisingDataLists:
        if advertisingDataList[3] not in ads:
            ads.append(advertisingDataList[3])
            if advertisingDataList[1] == 'SP':
                result = [advertisingDataList[3], float(advertisingDataList[2]), 0, 0, 0]
                resultAdvertisings.append(result)
            elif advertisingDataList[1] == 'SD':
                result = [advertisingDataList[3], 0, float(advertisingDataList[2]), 0, 0]
                resultAdvertisings.append(result)
            elif advertisingDataList[1] == 'SB':
                result = [advertisingDataList[3], 0, 0, float(advertisingDataList[2]), 0]
                resultAdvertisings.append(result)
            elif advertisingDataList[1] == 'SBV':
                result = [advertisingDataList[3], 0, 0, 0, float(advertisingDataList[2])]
                resultAdvertisings.append(result)
        else:
            for resultAdvertising in resultAdvertisings:
                if resultAdvertising[0] == advertisingDataList[3]:
                    if advertisingDataList[1] == 'SP':
                        resultAdvertising[1] += float(advertisingDataList[2])
                    elif advertisingDataList[1] == 'SD':
                        resultAdvertising[2] += float(advertisingDataList[2])
                    elif advertisingDataList[1] == 'SB':
                        resultAdvertising[3] += float(advertisingDataList[2])
                    elif advertisingDataList[1] == 'SBV':
                        resultAdvertising[4] += float(advertisingDataList[2])
    return resultAdvertisings


def exchangeRate():
    exchangeRateLists = []
    for row in workingSheetHL.iter_rows(min_row = 2, max_col = 2):
        exchangeRateList = [row[0].value, row[1].value]
        exchangeRateLists.append(exchangeRateList)
    return exchangeRateLists


def writeExcel(businessDatas, campaignsDatas, exchangeRateLists):
    productDatas = []
    costDatas = []
    for row in workingSheetSJ.iter_rows(min_row=2, max_col=11):
        row0 = row[0].value
        row1 = row[1].value
        row2 = row[2].value
        row3 = row[3].value
        row4 = row[4].value
        row5 = row[5].value
        row6 = row[6].value
        row7 = row[7].value
        row8 = row[8].value
        row9 = row[9].value
        row10 = row[10].value
        dataList1 = [row0, row1, row2, row3, row4, row5, row6]
        dataList2 = [row5, row7, row8, row9, row10]
        workingSheet.append(dataList1)
        productDatas.append(dataList1)
        costDatas.append(dataList2)

    for rowData in workingSheet.iter_rows(min_row=2, max_col=25):
        for businessData in businessDatas:
            if rowData[0].value == businessData[0] and rowData[6].value == businessData[1]:
                rowData[7].value = float(businessData[4])
                rowData[8].value = int(businessData[3])
                rowData[9].value = int(businessData[2])
        if rowData[8].value is not None and rowData[9].value is not None:
            rowData[10].value = rowData[8].value / rowData[9].value
        else:
            rowData[10].value = 0
        rowData[10].number_format = '0.00%'
        for campaignsData in campaignsDatas:
            if rowData[5].value == campaignsData[0]:
                rowData[11].value = campaignsData[1]
                rowData[12].value = campaignsData[2]
                rowData[13].value = campaignsData[3]
                rowData[14].value = campaignsData[4]
                rowData[15].value = campaignsData[1] + campaignsData[2] + campaignsData[3] + campaignsData[4]

        if rowData[7].value == None:
            rowData[7].value = 0
        if rowData[8].value == None:
            rowData[8].value = 0
        if rowData[9].value == None:
            rowData[9].value = 0
        if rowData[10].value == None:
            rowData[10].value = 0
        if rowData[11].value == None:
            rowData[11].value = 0
        if rowData[12].value == None:
            rowData[12].value = 0
        if rowData[13].value == None:
            rowData[13].value = 0
        if rowData[14].value == None:
            rowData[14].value = 0
        if rowData[15].value == None:
            rowData[15].value = 0

        if rowData[15].value != 0 and rowData[7].value != 0:
            rowData[16].value = rowData[15].value / rowData[7].value
        else:
            rowData[16].value = 0
        rowData[16].number_format = '0.00%'
        rowData[17].value = 0

        for costData in costDatas:
            if rowData[5].value == costData[0]:
                rowData[18].value = costData[1]
                rowData[19].value = costData[2]
                rowData[20].value = costData[3]
                rowData[22].value = costData[4]
        rowData[21].value = rowData[7].value * 0.15
        if rowData[7].value == 0 and rowData[15].value == 0:
            rowData[23].value = 0
        else:
            rowData[23].value = rowData[7].value - rowData[15].value - (rowData[18].value * rowData[8].value / 6.4) - (rowData[
                19].value * rowData[8].value / 6.4) -(rowData[20].value * rowData[8].value) - rowData[21].value - rowData[22].value
        if rowData[23].value != 0 and rowData[7].value != 0:
            rowData[24].value = rowData[23].value / rowData[7].value
        else:
            rowData[24].value = 0
        rowData[24].number_format = '0.00%'

        #汇率换算
        for exchangeRateList in exchangeRateLists:
            if exchangeRateList[0] == rowData[0].value:
                rowData[7].value = round(rowData[7].value * exchangeRateList[1], 2 )
                rowData[11].value = round(rowData[11].value * exchangeRateList[1], 2)
                rowData[12].value = round(rowData[12].value * exchangeRateList[1], 2)
                rowData[13].value = round(rowData[13].value * exchangeRateList[1], 2)
                rowData[14].value = round(rowData[14].value * exchangeRateList[1], 2)
                rowData[15].value = round(rowData[15].value * exchangeRateList[1], 2)
                rowData[18].value = round(rowData[18].value * 1)
                rowData[19].value = round(rowData[19].value * 1)
                rowData[20].value = round(rowData[20].value * exchangeRateList[1], 2)
                rowData[21].value = round(rowData[21].value * exchangeRateList[1], 2)
                rowData[22].value = round(rowData[22].value * exchangeRateList[1], 2)
                rowData[23].value = round(rowData[23].value * exchangeRateList[1] * 6.4, 2)


def bian():
    bian = Side(border_style='thin', color='000000')
    border = Border(top = bian, bottom = bian, left = bian, right = bian)
    rows = workingSheet.max_row
    cols = workingSheet.max_column
    for r in range(1 , rows + 1):
        for c in range(1, cols + 1):
            workingSheet.cell(r, c).border = border

def main():
    # 获取数据文件路径
    fileList = folder_name()
    businessDatas = business(fileList)
    campaignsDatas = campaigns(fileList)
    exchangeRateLists = exchangeRate()
    writeExcel(businessDatas, campaignsDatas, exchangeRateLists)
    bian()
    print('写入完成')


if __name__ == '__main__':
    workExcelPath = ('.//data/report/report.xlsx')
    # 读取工作表
    workingbook: Workbook = openpyxl.load_workbook(filename=workExcelPath, data_only=True)
    # 加载输出操作表
    workingSheet: Worksheet = workingbook['Sheet1']
    workingSheetGG: Worksheet = workingbook['广告名称对应ASIN']
    workingSheetSJ: Worksheet = workingbook['产品数据']
    workingSheetHL:workingSheet = workingbook['汇率']
    main()
    workingbook.save(filename=workExcelPath)
    ```
